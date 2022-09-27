import pandas as pd
import re as re
import datetime
from openpyxl import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import *
from openpyxl.worksheet import page
from GetDatabase import get_assoconnect_data_base
import math
import toml

liste_profs = toml.load("parameters.toml")["profs"]
liste_couleurs = toml.load("parameters.toml")["couleurs"]
annee_scolaire = "22-23"

#workbook = get_assoconnect_data_base()


workbook = pd.read_excel("export.xlsx")


class NiveauDanse():

    def __init__(self, nom_niveau: str):
        self.nom_niveau = nom_niveau
        self.discipline = self.def_discipline()
        self.liste_cours = []

    def def_discipline(self) -> str:
        """
        Détermine la discipline à partir du titre du niveau
        :return: le nom de la discipline (str)
        """
        if re.search("Classique", self.nom_niveau):
            return "Classique"
        elif re.search("Modern Jazz", self.nom_niveau):
            return "Jazz"
        elif re.search("Contemporain", self.nom_niveau):
            return "Contemporain"
        elif re.search("Caractère", self.nom_niveau):
            return "Caractère"
        elif re.search("Eveil", self.nom_niveau):
            return "Eveil/Initiation"
        elif re.search("Initiation", self.nom_niveau):
            return "Eveil/Initiation"
        # todo: vérifier que baroque et barre au sol apparaissent tel quel dans les bases de données adulte…
        elif re.search("Baroque", self.nom_niveau):
            return "Baroque"
        elif re.search("Barre au Sol", self.nom_niveau):
            return "Barre au Sol"

    def __repr__(self):
        return self.nom_niveau + " " + self.discipline + " " + str(self.liste_cours)


class CoursDanse():
    def __init__(self, nom_cours: str, discipline: str = ""):
        self.nom_cours = nom_cours
        self.jour = self.def_jour()
        self.heure = self.def_heure()
        self.prof = self.def_prof()
        self.liste_eleves = []
        self.nb_eleves = 0
        self.df = {}
        self.discipline = discipline

    def __repr__(self):
        return str(self.jour) + str(self.heure) + str(self.prof)

    def def_jour(self):
        jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        for jour in jours:
            if re.search(jour, self.nom_cours):
                return jour

    # retourne l'heure sous la forme 15h30
    def def_heure(self):
        try:
            pattern = re.compile("[0-9][0-9]h[0-9][0-9]")
            return pattern.findall(self.nom_cours)[0]
        except IndexError:
            pattern = re.compile("[0-9]h[0-9][0-9]")
            return "0" + pattern.findall(self.nom_cours)[0]

    def def_prof(self):
        for prof in liste_profs:
            if re.search(liste_profs[prof]["nom"], self.nom_cours):
                return liste_profs[prof]
            elif re.search(liste_profs[prof]["diminutif"], self.nom_cours):
                return liste_profs[prof]


def chercher_cours(niveau):
    liste_noms_cours = []
    discipline = niveau.discipline
    # on cherche tous les cours du niveau et on le met dans la liste liste_cours
    for index in workbook.index:
        if workbook.loc[index, niveau.nom_niveau] != 0 and type(workbook.loc[index, niveau.nom_niveau]) == str:
            cours = workbook.loc[index, niveau.nom_niveau].split("|")
            for i in cours:
                if i not in liste_noms_cours:
                    liste_noms_cours.append(i)
    # on fait de chaque élément de la liste un élément de la classe CoursDanse
    liste_cours = []
    for i in liste_noms_cours:
        liste_cours.append(CoursDanse(i, discipline))
    return liste_cours


# DÉBUT DU PROCESS


# On recherche tous les niveaux de cours
liste_niveaux = []
for col in workbook:
    if re.search("Cours", col) and re.search("saison {}".format(annee_scolaire), col):
        if col == "Cours de danse choisi(s)" or col == "Cours d'essai":
            pass
        else:
            liste_niveaux.append(NiveauDanse(nom_niveau=col))

# on génère la liste de cours pour chaque niveau de cours (niveau étant élément de liste_niveau de class NiveauDanse)
for niveau in liste_niveaux:
    niveau.liste_cours = chercher_cours(niveau)

# pour chaque cours on recherche les élèves
for niveau in liste_niveaux:
    for cours in niveau.liste_cours:
        liste_eleves = []
        # on recherche dans toutes les cases non vides du tableau les cours
        for index, row in workbook.iterrows():
            case = workbook[niveau.nom_niveau][index]
            if case != 0 and case != "":
                # si la case correspond au cours
                for i in str(case).split("|"):
                    if i == cours.nom_cours:
                        # on recherche tous les cours que fait l'élève en dehors du cours en cours
                        list_autres_cours = []
                        for niv in liste_niveaux:
                            if row[niv.nom_niveau] != 0:
                                for e in str(row[niv.nom_niveau]).split("|"):
                                    if e != cours.nom_cours and e != "nan":
                                        list_autres_cours.append(e)
                                autres_cours = ""
                                for i in list_autres_cours:
                                    autres_cours += i + " | "

                        # on détermine l'âge de l'élève à partir de sa date de naissance
                        try:
                            age = int((datetime.datetime.now() - row["Date de naissance"]).days / 365)
                        except TypeError:
                            print("Age non trouvé pour {} {}".format(row["Nom"], row["Prénom"]))
                            age = "?"
                        except ValueError:
                            print("Age non trouvé pour {} {}".format(row["Nom"], row["Prénom"]))
                            age = "?"

                        # telephone
                        telephone = str(row["Téléphone mobile"])
                        if len(telephone) >= 10:
                            telephone = "0" + telephone[2] + " " + telephone[3] + telephone[4] + " " + telephone[5] + \
                                        telephone[6] + " " + telephone[7] + telephone[8] + " " + telephone[9] + \
                                        telephone[10]

                        eleve = {"ID": row["ID du Contact"],
                                 "Nom": row["Nom"],
                                 "Prénom": row["Prénom"],
                                 "Âge": age,
                                 "Genre": row["Sexe"],
                                 "Téléphone": telephone,
                                 "Mail": row["Email"],
                                 "Autres cours": autres_cours}
                        # on ne prends en compte que les élèves vérifiés.
                        if row["Inscription 2022-2023 vérifiée"] == "oui":
                            liste_eleves.append(eleve)
                        else:
                            print(f"{eleve['Nom']} {eleve['Prénom']} a son inscription 2022/23 non vérifiée.")
        # on trie la liste des élèves par prénoms
        liste_eleves.sort(key=lambda x: x["Prénom"])
        # on ajoute la liste des élèves à l'instance de la classe CoursDanse
        cours.liste_eleves = liste_eleves
        # on ajoute le nb d'élèves à l'instance de la classe CoursDanse
        cours.nb_eleves = len(cours.liste_eleves)


def add_value(list, key, eleve):
    list.append(eleve.get(key))


for niveau in liste_niveaux:
    for cours in niveau.liste_cours:
        df = {}
        prenom, nom, age, genre, autres_cours, telephone, mail, id = [], [], [], [], [], [], [], []
        for eleve in cours.liste_eleves:
            add_value(prenom, "Prénom", eleve)
            add_value(nom, "Nom", eleve)
            add_value(age, "Âge", eleve)
            add_value(genre, "Genre", eleve)
            add_value(autres_cours, "Autres cours", eleve)
            add_value(telephone, "Téléphone", eleve)
            add_value(mail, "Mail", eleve)
            add_value(id, "ID", eleve)

        df["Prénom"] = prenom
        df["Nom"] = nom
        df["Âge"] = age
        df["Genre"] = genre
        df["Autres cours"] = autres_cours
        df["Téléphone"] = telephone
        df["Mail"] = mail
        df["ID"] = id
        # ajout de la df du cours à l'instance de la classe CoursDanse
        df = pd.DataFrame(df)
        cours.df = df

# tri des cours dans la semaine
semaine = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
lundi, mardi, mercredi, jeudi, vendredi, samedi, dimanche = [], [], [], [], [], [], []
var_semaine = [lundi, mardi, mercredi, jeudi, vendredi, samedi, dimanche]
for niveau in liste_niveaux:
    for cours in niveau.liste_cours:
        for jour in semaine:
            if cours.jour == jour:
                # on évite les doublons
                if cours not in var_semaine[semaine.index(jour)]:
                    var_semaine[semaine.index(jour)].append(cours)

for jour in var_semaine:
    jour.sort(reverse=False, key=lambda x: (x.heure, x.nom_cours))


# GÉNÉRATION DU TABLEAU D'APPEL


def contsruction_tableau_appel():
    # écriture sur le fichier excel

    wb = Workbook()
    ws = wb.active
    ws.title = "Lundi"
    wb.create_sheet("Mardi")
    wb.create_sheet("Mercredi")
    wb.create_sheet("Jeudi")
    wb.create_sheet("Vendredi")
    wb.create_sheet("Samedi")
    wb.create_sheet("Dimanche")

    # définition des styles
    row1 = NamedStyle(name="row1")
    row1.font = Font(name="Arial", size=10)
    row1.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])

    row2 = NamedStyle(name="row2")
    row2.font = Font(name="Arial", size=10)
    row2.fill = PatternFill("solid", start_color=liste_couleurs["blanc"])

    sub = NamedStyle(name="sub")
    sub.font = Font(name="Arial", size=11, color=liste_couleurs["blanc"])
    sub.fill = PatternFill("solid", start_color=liste_couleurs["gris"])

    prof_style = NamedStyle(name="prof")

    def mise_en_forme(cell, style):
        # applique le style à la cellule
        cell.style = style

    def mise_en_forme_prof(cell, prof):
        # change la couleur en fonction du prof
        if prof is None:
            prof = liste_profs["none"]
        prof_style.font = Font(name="Arial", size=13, color=liste_couleurs[prof["couleur"]])
        prof_style.fill = PatternFill("solid", start_color=liste_couleurs[prof["fond"]])
        mise_en_forme(cell, prof_style)

    for jour in semaine:
        ws = wb[jour]
        for cours in var_semaine[semaine.index(jour)]:
            if cours.prof is None:
                print("Impossible de déterminer la·le prof du cours : {}".format(cours.nom_cours))
            # titre du tableau
            ws.append([cours.nom_cours, "", "", "", "", cours.discipline, "Total élèves : " + str(cours.nb_eleves)])
            ws.merge_cells(start_row=ws.max_row, end_row=ws.max_row, start_column=1, end_column=5)
            ws.merge_cells(start_row=ws.max_row, end_row=ws.max_row, start_column=7, end_column=8)
            for cell in ws[ws.max_row]:
                mise_en_forme_prof(cell, cours.prof)

            # corps du tableau avec entêtes
            compteur = 1
            for row in dataframe_to_rows(cours.df, index=False, header=True):
                ws.append(row)
                # mise en forme

                for cell in ws[ws.max_row]:
                    if compteur == 1:
                        mise_en_forme(cell, sub)
                    elif compteur == 2:
                        mise_en_forme(cell, row2)
                    elif compteur == 3:
                        mise_en_forme(cell, row1)
                if compteur == 1:
                    compteur = 2
                elif compteur == 2:
                    compteur = 3
                elif compteur == 3:
                    compteur = 2
            ws.append([" ", " ", " ", " ", " "])
    for sheet in wb:
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 5
        sheet.column_dimensions["D"].width = 8
        sheet.column_dimensions["E"].width = 25
        sheet.column_dimensions["F"].width = 15
        sheet.column_dimensions["G"].width = 19
        sheet.column_dimensions["H"].width = 8

    # sauvegarde du fichier xlsx de sortie
    print("génération du tableau d'appel terminée")
    wb.save("liste_appel.xlsx")


# GÉNÉRATION DU TABLEAU D'ÉTIQUETTES


def construction_tableau_etiquettes():
    def eleve_etiquette(cours):
        """"
        retourne une liste d'élève dans laquelle chaque élève est un dictionnaire
        """
        liste_eleve_cours = []
        # pour chaque élève on fait un dictionnaire qu'on ajoute dans la liste des élèves du cours
        for index, ligne in cours.df.iterrows():
            eleve = {}
            eleve["Cours"] = cours.discipline[:4] + " " + cours.jour[:2] + \
                             " " + cours.heure + " " + liste_profs[cours.prof][2]
            eleve["Couleur"] = liste_profs[cours.prof][3]
            eleve["Nom"] = ligne["Nom"]
            eleve["Prénom"] = ligne["Prénom"]
            eleve["Âge"] = ligne["Âge"]
            eleve["Autres cours"] = ligne["Autres cours"].split(" | ")
            # on renomme les autres cours : Co Ve 17h30 Nat

            for i in eleve["Autres cours"]:
                if i == "" or math.isnan(i) or i == "nan":
                    eleve["Autres cours"].remove(i)
                for niveau in liste_niveaux:
                    for coursname in niveau.liste_cours:
                        if coursname.nom_cours == i:
                            eleve["Autres cours"][eleve["Autres cours"].index(i)] = coursname.discipline[
                                                                                    :4] + " " + coursname.jour[:2] + \
                                                                                    " " + coursname.heure + " " + \
                                                                                    liste_profs[coursname.prof][2]
            # On enlève les cours dans Autres cours qui sont en double
            autres_cours_sans_doublon = []
            for _autre_cours in eleve["Autres cours"]:
                if _autre_cours not in autres_cours_sans_doublon:
                    autres_cours_sans_doublon.append(_autre_cours)
            eleve["Autres cours"] = autres_cours_sans_doublon
            liste_eleve_cours.append(eleve)
        liste_eleve_cours = index_eleves(liste_eleve_cours)
        # on retourne la liste
        return liste_eleve_cours

    def index_eleves(liste_eleve_cours):
        # on ajoute l'index de chaque élève
        liste_eleve_cours.sort(key=lambda x: x["Prénom"])
        for i in liste_eleve_cours:
            i["Index"] = liste_eleve_cours.index(i) + 1
        # on retourne la liste
        return liste_eleve_cours

    # lister tous les élèves de chaque cours en prenant en compte les cours en doublons et
    # en les fusionnant pour qu'aucun élève ne soit oublié car il manque des élèves dans certains doublons…
    liste_index = []
    liste_eleves_cours = []
    for niveau in liste_niveaux:
        for cours in niveau.liste_cours:
            # on fabrique un identifiant pour le cours pour trouver les doublons
            code = cours.discipline + cours.prof + cours.jour + cours.heure
            if code not in liste_index:
                liste_index.append(code)
                # on génère la liste des élèves dans le cours (un élève: dictionnaire)
                liste_eleves_cours.append(eleve_etiquette(cours))
            else:
                # on génère la liste des élèves dans le cours en doublon pour pouvoir ensuite la comparer avec la
                # liste originale
                doublon = eleve_etiquette(cours)
                # on comparer la liste d'origine avec la liste doublon
                for eleve in liste_eleves_cours[liste_index.index(code)]:
                    for eleve_doublon in doublon:
                        # on retire de la liste des doublons les élèves qui se trouvent dans la liste originale
                        if eleve["Nom"] == eleve_doublon["Nom"] and eleve["Prénom"] == eleve_doublon["Prénom"]:
                            doublon.remove(eleve_doublon)
                # on ajoute les élèves restant s'il y en a à la liste originale de sorte à ce qu'il y ait tous les
                # élèves dans la liste originale
                for eleve_restant in doublon:
                    liste_eleves_cours[liste_index.index(code)].append(eleve_restant)
                    # on réactualise l'index des élèves
                    liste_eleves_cours[liste_index.index(code)] = index_eleves(
                        liste_eleves_cours[liste_index.index(code)])

    # Pour chaque élève on enlève le cours actuel dans la liste Autres cours
    for cours in liste_eleves_cours:
        for eleve in cours:
            for autre_cours in eleve["Autres cours"]:
                if autre_cours == eleve["Cours"]:
                    eleve["Autres cours"].remove(eleve["Cours"])

    # Construction du tableau
    wb = Workbook()
    # on enlève la feuille par défaut
    wb.remove(wb['Sheet'])

    # styles
    # bordure
    bordure = Side(border_style="thick", color="000000")
    # prenom
    style_prenom = NamedStyle(name="style_prenom")
    style_prenom.font = Font(name="Arial", size=16, bold=True)
    style_prenom.alignment = Alignment(horizontal="right", vertical="center")
    # nom
    style_nom = NamedStyle(name="style_nom")
    style_nom.font = Font(name="Arial", size=12)
    style_nom.alignment = Alignment(horizontal="center", vertical="bottom")
    # titre
    style_titre = NamedStyle(name="style_titre")
    style_titre.font = Font(name="Arial", size=14, bold=True, underline="single")
    style_titre.alignment = Alignment(horizontal="center", vertical="bottom")
    # index
    style_index = NamedStyle(name="style_index")
    style_index.font = Font(name="Arial", size=14, bold=True)
    style_index.alignment = Alignment(horizontal="left", vertical="center")
    # autres cours
    style_autres_cours = NamedStyle(name="style_autres_cours")
    style_autres_cours.font = Font(name="Arial", size=12)
    # couleur / commentaire
    style_couleur = NamedStyle(name="style_couleur")
    style_couleur.font = Font(name="Arial", size=12, bold=True)
    style_couleur.alignment = Alignment(horizontal="right", vertical="center")
    style_couleur.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])
    # commentaire
    style_commentaire = NamedStyle(name="style_commentaire")
    style_commentaire.font = Font(name="Arial", size=11)
    style_commentaire.alignment = Alignment(horizontal="center")
    style_commentaire.fill = PatternFill("solid", start_color=liste_couleurs["grisClair"])
    # age
    style_age = NamedStyle(name="style_age")
    style_age.font = Font(name="Arial", size=12)
    style_age.alignment = Alignment(horizontal="right", vertical="center")

    eleve_vide = {"Cours": " ", "Couleur": " ", "Nom": " ", "Prénom": " ", "Âge": " ", "Autres cours": " ",
                  "Index": " "}

    # on tri les cours par jour et par heure…
    days = {"Lu": 1, "Ma": 2, "Me": 3, "Je": 4, "Ve": 5, "Sa": 6, "Di": 7}
    liste_eleves_cours.sort(reverse=False, key=lambda x: (days[x[0]["Cours"][5:7]], x[0]["Cours"][8:13]))

    # on ajoute des élèves blancs dans chaque cours (3 dans les cours où le nb d'élève est impaire et 2 dans le
    # cours où le nb d'élève est pair de sorte à faire un espace vide entre chaque cours
    for cours in liste_eleves_cours:
        if len(cours) % 2 == 0:
            cours.append(eleve_vide)
            cours.append(eleve_vide)
        else:
            cours.append(eleve_vide)
            cours.append(eleve_vide)
            cours.append(eleve_vide)

    # on détermine les différents messages imprimés dans les différentes pages
    messages = ["Étiquette plastifiée à laisser sur le sac",
                "À donner par l'enfant en arrivant en répétition",
                "À garder pour récupérer l’enfant après la répétition",
                "À donner par l'enfant en arrivant au spectacle",
                "Étiquette Costume Gala"]
    messages_a_imprimer = [["Étiquette plastifiée à laisser", "sur le sac"],
                           ["À donner par l'enfant en arrivant", "en répétition"],
                           ["À garder pour récupérer l’enfant", "après la répétition"],
                           ["À donner par l'enfant en arrivant", "au spectacle"],
                           ["Étiquette Costume Gala", ""]]

    for message in messages:
        # on créé une nouvelle page en fonction du message à afficher sur les étiquettes
        print("début impression de la page {}.".format(str(messages.index(message) + 1)))
        ws = wb.create_sheet(str(messages.index(message) + 1))
        ws.title = str(messages.index(message) + 1)

        compteur = 0

        for cours in liste_eleves_cours:
            # on affiche deux élèves à la fois
            col = 1
            for eleve in cours:
                try:
                    eleve2 = cours[cours.index(eleve) + 1]
                except IndexError:
                    pass
                if col == 1:
                    if compteur == 6:
                        compteur = 0

                    # nom prénom
                    ws.append([eleve["Nom"], " ", eleve["Prénom"], " ", eleve2["Nom"], " ", eleve2["Prénom"]])

                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_nom, style_nom
                    ws[ws.max_row][2].style, ws[ws.max_row][6].style = style_prenom, style_prenom

                    # Cours
                    ws.append([eleve["Cours"], " ", " ", " ", eleve2["Cours"]])
                    ws[ws.max_row][0].style = style_titre
                    ws[ws.max_row][4].style = style_titre
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=8)
                    # Index / age
                    ws.append(["  " + str(eleve["Index"]), " ", str(eleve["Âge"]) + " ans", " ",
                               "  " + str(eleve2["Index"]), " ", str(eleve2["Âge"]) + " ans"])
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_index, style_index
                    ws[ws.max_row][2].style, ws[ws.max_row][6].style = style_age, style_age
                    # Autres cours
                    # premier rang
                    impr = []
                    espace = " "
                    if len(eleve["Autres cours"]) == 0:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) == 1:
                        impr.append(espace + eleve["Autres cours"][0])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) == 2:
                        impr.append(espace + eleve["Autres cours"][0])
                        impr.append(" ")
                        impr.append(eleve["Autres cours"][1])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    if len(eleve2["Autres cours"]) == 0:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) == 1:
                        impr.append(espace + eleve2["Autres cours"][0])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) == 2:
                        impr.append(espace + eleve2["Autres cours"][0])
                        impr.append(" ")
                        impr.append(eleve2["Autres cours"][1])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    ws.append(impr)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][2].style = style_autres_cours, style_autres_cours
                    ws[ws.max_row][4].style, ws[ws.max_row][6].style = style_autres_cours, style_autres_cours

                    # deuxième rang
                    impr = []
                    if len(eleve["Autres cours"]) == 3:
                        impr.append(eleve["Autres cours"][2])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve["Autres cours"]) >= 4:
                        impr.append(eleve["Autres cours"][2])
                        impr.append(" ")
                        impr.append(eleve["Autres cours"][3])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    if len(eleve2["Autres cours"]) == 3:
                        impr.append(eleve2["Autres cours"][2])
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    elif len(eleve2["Autres cours"]) >= 4:
                        impr.append(eleve2["Autres cours"][2])
                        impr.append(" ")
                        impr.append(eleve2["Autres cours"][3])
                        impr.append(" ")
                    else:
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                        impr.append(" ")
                    ws.append(impr)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=2)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=3, end_column=4)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=6)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=7, end_column=8)
                    ws[ws.max_row][0].style, ws[ws.max_row][2].style = style_autres_cours, style_autres_cours
                    ws[ws.max_row][4].style, ws[ws.max_row][6].style = style_autres_cours, style_autres_cours

                    # couleur + message
                    ws.append([messages_a_imprimer[messages.index(message)][0], " ", " ", eleve["Couleur"],
                               messages_a_imprimer[messages.index(message)][0], " ", " ", eleve2["Couleur"]])
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=3)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=7)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_commentaire, style_commentaire
                    ws[ws.max_row][3].style, ws[ws.max_row][7].style = style_couleur, style_couleur
                    # deuxième ligne commentaire
                    ws.append([messages_a_imprimer[messages.index(message)][1], " ", " ", " ",
                               messages_a_imprimer[messages.index(message)][1], " ", " "])
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=1, end_column=3)
                    ws.merge_cells(start_row=int(ws.max_row), end_row=int(ws.max_row),
                                   start_column=5, end_column=7)
                    ws[ws.max_row][0].style, ws[ws.max_row][4].style = style_commentaire, style_commentaire
                    ws[ws.max_row][3].style, ws[ws.max_row][7].style = style_commentaire, style_commentaire
                    col = 2
                    compteur += 1
                else:
                    col = 1

        # ajout des bordures
        # verticales
        for row in ws:
            row[0].border = Border(left=bordure)
            row[3].border = Border(right=bordure)
            row[4].border = Border(left=bordure)
            row[7].border = Border(right=bordure)
        # horizontales
        compteur = 0
        for row in ws:
            compteur += 1
            if compteur == 1:
                row[0].border = Border(top=bordure, left=bordure)
                row[1].border = Border(top=bordure)
                row[2].border = Border(top=bordure)
                row[3].border = Border(top=bordure, right=bordure)
                row[4].border = Border(top=bordure, left=bordure)
                row[5].border = Border(top=bordure)
                row[6].border = Border(top=bordure)
                row[7].border = Border(top=bordure, right=bordure)
            elif compteur == 7:
                row[0].border = Border(bottom=bordure, left=bordure)
                row[1].border = Border(bottom=bordure)
                row[2].border = Border(bottom=bordure)
                row[3].border = Border(bottom=bordure, right=bordure)
                row[4].border = Border(bottom=bordure, left=bordure)
                row[5].border = Border(bottom=bordure)
                row[6].border = Border(bottom=bordure)
                row[7].border = Border(bottom=bordure, right=bordure)
                compteur = 0
        # définition de la larguer des colonnes
        ws.column_dimensions["A"].width = 11.90
        ws.column_dimensions["B"].width = 11.90
        ws.column_dimensions["C"].width = 11.90
        ws.column_dimensions["D"].width = 11.90
        ws.column_dimensions["E"].width = 11.90
        ws.column_dimensions["F"].width = 11.90
        ws.column_dimensions["G"].width = 11.90
        ws.column_dimensions["H"].width = 11.90

        # définition de la zone d'impression et des marges
        ws.page_margins = page.PageMargins(left=0.23, right=0.23, top=0.23, bottom=0.23, header=0.23, footer=0.23)
        wb.page_margins = page.PageMargins(left=0.23, right=0.23, top=0.23, bottom=0.23, header=0.23, footer=0.23)
        print("Impression de la page {} terminée".format(str(messages.index(message) + 1)))
    # sauvegarde du fichier
    wb.save("Étiquettes.xlsx")




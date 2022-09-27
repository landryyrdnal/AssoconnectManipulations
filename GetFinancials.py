from time import sleep

import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import os
import toml
import re
from openpyxl import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import *
from operator import itemgetter

rouge = "FF0000"
bleu = "00B0F0"
vert = "92D050"
violet = "7030A0"

def GetFinancialDb(excel_db:str)-> pd.DataFrame:
    workbook = pd.read_excel(excel_db)
    return workbook

def AssoConnectConnection()-> webdriver.chrome:
    """
    Cette fonction va chercher sur AssoConnect la dernière mise à jour de toutes les données et retourne un dataframe.
    :return: un dataframe Pandas avec toutes les données d'AssoConnect actualisées
    """

    params = toml.load("./parameters.toml")
    password = params["AssoConnectFinancial"]["password"]
    mail = params["AssoConnectFinancial"]["mail"]
    url = params["AssoConnect"]["connection_page"]
    current_dir = os.getcwd()

    chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-gpu")
    prefs = {'download.default_directory' : current_dir}
    chrome_options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(chrome_options=chrome_options, options=chrome_options)
    # direction vers la page
    driver.set_window_size(1900, 1060)
    driver.get(url)

    # entrée des codes & validation
    driver.find_element(By.XPATH, """/html/body/div/div[3]/div[1]/div[1]/form/div[1]/input""").send_keys(mail)
    driver.find_element(By.XPATH, """/html/body/div/div[3]/div[1]/div[1]/form/button/span""").click()

    driver.find_element(By.XPATH, """/html/body/div/div[3]/div[1]/div[1]/form/div[3]/input""").send_keys(password)
    driver.find_element(By.XPATH, """/html/body/div/div[3]/div[1]/div[1]/form/button/span""").click()



    try:
    # cliquer sur administration
        driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/div[2]/a/div/span").click()
        print("connection OK")
        print("entrée sur la page ADMINISTRATION OK")
    except:
        print("Impossible de se connecter")

    return driver


def GetFinancialInformation(db:pd.DataFrame, driver:webdriver.chrome)->pd.DataFrame:
    db_output = []
    for index in db.index:

        adresse = "https://academie-de-ballet-nini-theilade.assoconnect.com/" + db.loc[index, "Détails"].split("https:///")[1]
        eleve = {"Prénom":db.loc[index, "Prénom participant"],
                 "Nom":db.loc[index, "Nom participant"],
                 "URL": adresse,
                 "N° de transaction": db.loc[index, "N° de transaction"],
                 "Moyen de paiement": db.loc[index, "Moyen de paiement"],
                 "restant à payer": "?"}
        # si l'élève a payé en plusieurs fois
        ################# CONDITION AJOUTÉE POUR DEBUG
        print(adresse)
        if eleve["Moyen de paiement"] == "Paiement en plusieurs fois" or eleve["Moyen de paiement"] == "Chèque":
            continuer = False
            tentative = 0
            while continuer is False:
                try:
                    driver.get(adresse)
                    continuer = True
                except:
                    sleep(1)
                    tentative += 1
                    print("tentative de connection{}".format(tentative))


            # on cherche dans la page s'il y a eu un paiement
            for i in range(1, 8):
                try:
                    a = driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/section/div/div[3]/div[3]/div[2]/div[{}]".format(i))
                    if "Paiement en plusieurs fois" in a.text:
                        # on entre dans la page du détail du paiement en plusieurs fois
                        text_payment= []
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/section/div/div[3]/div[3]/div[2]/div[4]/div/div[2]/div/a/div/span".format(i)).click()
                        # on retente tant que la page n'a pas chargé correctement
                        while len(text_payment) < 2:
                            text_payment = driver.find_element(By.XPATH, "/html/body").text.split("ÉCHÉANCE")
                        echeances = []
                        # on met les trois échéances de la personne dans un tableau
                        for e in range(1,4):

                            paiement = text_payment[e]
                            date = re.findall("[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", paiement)[0]
                            statut = "???"
                            if "Planifié" in paiement:
                                statut = "Planifié"
                            elif "Payé" in paiement:
                                statut = "Payé"
                            elif "Abandonné" in paiement:
                                statut = "Abandonné"
                            montant = float(paiement.split("Montant :")[1].split("€")[0].replace(" ","").replace(",","."))
                            echeances.append({"date":date, "statut":statut, "montant":montant})
                        # on cherche le restant à payer

                        driver.get(adresse)

                        # on entre sur la page des écritures des opérations en ligne
                        driver.find_element(By.XPATH,"/html/body/div[1]/div[3]/section/div/div[3]/div[3]/div[2]/div[3]/div[2]").click()
                        continuer = False
                        while continuer is False:
                            texte = driver.find_element(By.XPATH, "/html/body").text
                            if "Paiements enregistrés : " in texte:

                                restant = texte.split("Paiements enregistrés : ")[1].split(" EUR")[0].replace(",",".").replace(" ","").replace("\u202f","")
                                restant = float(restant.split(" / ")[1]) - float(restant.split(" / ")[0])
                                eleve.update({"restant à payer": str(restant)})
                                continuer = True
                            elif "Reste à recevoir : " in texte:
                                restant = texte.split("Reste à recevoir :")[1].split("€")[0]
                                eleve.update({"restant à payer": str(restant)})
                                continuer = True

                        # on ajoute l'élève à la base de donnée
                        eleve.update({"échéances": echeances})
                        db_output.append(eleve)
                        print(eleve)
                        break

                    elif "Chèque" in a.text:
                        # on va sur la page du détail du paiement
                        driver.find_element(By.XPATH, "/html/body/div[1]/div[3]/section/div/div[3]/div[3]/div[2]/div[5]/div[2]").click()
                        continuer = True
                        while continuer is True:

                            texte = driver.find_element(By.XPATH, "/html/body").text
                            if "Reste à recevoir : " in texte:
                                restant_a_payer = texte.split("Reste à recevoir : ")[1].split("€")[0]
                                paiements = texte.split("Journal : ")
                                echeances = []

                                for i in paiements:


                                    if i.startswith("VT"):
                                        montant = i.split("VT\n")[1].split(" EUR")[0].replace(" ","").replace("\u202f","")
                                        statut = "Payé"
                                        date = paiements[paiements.index(i)-1]
                                        date = re.findall("[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", date)[-1]
                                        echeances.append({"date": date, "statut": statut, "montant": montant})

                                    elif i.startswith("OD"):
                                        montant = i.split("OD\n")[1].split(" EUR")[0].replace(" ","").replace("\u202f","")
                                        statut = "Planifié"
                                        if "paiement du " in i:
                                            date = i.split("paiement du ")[1]
                                            date = re.findall("[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]", date)[0]
                                            echeances.append({"date": date, "statut": statut, "montant": montant})
                                        #else:
                                            #statut = "Pas encaissé"
                                eleve.update({"échéances": echeances})
                                eleve.update({"restant à payer": restant_a_payer})
                                db_output.append(eleve)
                                print(eleve)
                                continuer = False
                        break
                except selenium.common.exceptions.NoSuchElementException:
                    print("non trouvé")
                    pass

    return pd.DataFrame(db_output)

def excel_format(db:pd.DataFrame)->None:
    wb = Workbook()
    ws = wb.active
    wb = Workbook()
    ws = wb.active
    ws.title = "Export compta"
    style_paye = NamedStyle(name="style_paye")
    style_paye.fill = PatternFill("solid", start_color=vert)
    style_attente = NamedStyle(name="style_attente")
    style_attente.fill = PatternFill("solid", start_color=bleu)
    style_echec = NamedStyle(name="style_echec")
    style_echec.fill = PatternFill("solid", start_color=rouge)
    style_pas_encaisse = NamedStyle(name="style_pas_encaissé")
    style_echec.fill = PatternFill("solid", start_color=violet)

    entete = ["Prénom", "Nom","URL","N° de transaction","Moyen de paiement", "Reste à recevoir (les chèques non encaissés ne sont pas comptés)"]
    # on liste tous les mois des échéances
    dates = []
    # on liste tous les mois où il y a au moins une échéance
    for raw in db.index:
        echeances = db["échéances"][raw]
        for i in echeances:
            date = i["date"]

            if date not in dates:
                dates.append(date)

    index_dates = []
    for date in dates:
        index_dates.append(
            {"raw_date": date, "reversed_date": date.split("/")[2] + date.split("/")[1] + date.split("/")[0]})

    index_dates = sorted(index_dates, key=itemgetter('reversed_date'))
    dates = []
    for i in index_dates:
        dates.append(i["raw_date"])

    entete.extend(dates)
    ws.append(entete)
    # création d'une liste vide pour la remplir avec les données de paiement en fonction des mois
    empty_liste = []
    for i in dates:
        empty_liste.append("")

    # pour chaque élève on génère deux liste, la liste de mise en forme et la liste d'échéance
    for raw in db.index:
        echeances = db["échéances"][raw]
        tableau_echeance = empty_liste.copy()
        tableau_mise_en_forme = empty_liste.copy()
        for i in echeances:
            date = i["date"]
            for e in dates:
                if date == e:
                    index = dates.index(e)
                    tableau_echeance[index] = i["montant"]
                    tableau_mise_en_forme[index] = i["statut"]

        eleve = [db["Prénom"][raw],
                  db["Nom"][raw],
                  db["URL"][raw],
                  db["N° de transaction"][raw],
                  db["Moyen de paiement"][raw],
                  db["restant à payer"][raw]]
        eleve.extend(tableau_echeance)
        ws.append(eleve)

        # application des styles sur le tableau en fonction de l'état du paiement
        for style in tableau_mise_en_forme:
            if style == "Payé":
                ws[ws.max_row][tableau_mise_en_forme.index(style)+6].style = style_paye
            elif style == "Planifié":
                ws[ws.max_row][tableau_mise_en_forme.index(style)+6].style = style_attente
            elif style == "Abandonné":
                ws[ws.max_row][tableau_mise_en_forme.index(style) + 6].style = style_echec
            elif style == "Pas encaissé":
                ws[ws.max_row][tableau_mise_en_forme.index(style) + 6].style = style_pas_encaisse
            tableau_mise_en_forme[tableau_mise_en_forme.index(style)] = ""



        # mise en forme des liens
        for i in range(2, ws.max_row + 1):
            cell_url = ws.cell(row=i, column=3)
            cell_url.style = "Hyperlink"

    wb.save("results.xlsx")

db = GetFinancialDb("./exportcompta.xlsx")
driver = AssoConnectConnection()
db_output = GetFinancialInformation(db, driver)
driver.close()
db_for_excel = excel_format(db_output)
print("écriture de la base de donnée")




